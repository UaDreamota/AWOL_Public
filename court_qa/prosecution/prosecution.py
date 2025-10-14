# extract_awol_407_408_monthly.py
# pip install pandas openpyxl xlrd

import re
import sys
import csv
import traceback
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Tuple, List

import pandas as pd
from openpyxl import load_workbook

# ── CONFIG ─────────────────────────────────────────────────────────────────────
INPUT_DIR = r"D:\Projects\GitRepositories\TheAWOLthing\AWOL_Research\court_qa\prosecution"
SHEET_INDEX = 6  # 0-based → the 7th sheet as requested
ROW_407 = 115
ROW_408 = 116
COL_I = "I"  # value column
LABEL_COL = "B"  # top-left cell of merged B:G
OUTPUT_CUMULATIVE = "awol_cumulative_by_file.csv"
OUTPUT_MONTHLY = "awol_monthly_series.csv"

# For sanity: title should exist and look like a “Form” summary
EXPECTED_TITLE_HINTS = ["форма", "форма 1", "1-"]

# Month dictionaries (both Cyrillic and common translits in filenames/headings)
MONTHS_UA = {
    "січ": 1, "січень": 1, "сiч": 1,
    "лютий": 2, "лют": 2,
    "берез": 3, "березень": 3,
    "квіт": 4, "квітень": 4, "квiт": 4,
    "трав": 5, "травень": 5,
    "черв": 6, "червень": 6,
    "лип": 7, "липень": 7,
    "серп": 8, "серпень": 8,
    "верес": 9, "вересень": 9,
    "жовт": 10, "жовтень": 10,
    "листоп": 11, "листопад": 11,
    "груд": 12, "грудень": 12,
}
MONTHS_LAT = {
    "sichen":1, "sich":1,
    "lyut":2, "liuty":2, "lyutiy":2, "lyuty":2,
    "berez":3, "berezen":3,
    "kvit":4, "kviten":4,
    "trav":5, "traven":5,
    "cherv":6, "cherven":6,
    "lyp":7, "lypen":7, "lipen":7,
    "serp":8, "serpen":8,
    "veres":9, "veresen":9,
    "zhovt":10, "jovt":10, "zhovten":10, "jovten":10,
    "list":11, "listopad":11, "lystopad":11,
    "grud":12, "hrud":12, "gruden":12, "hruden":12,
}
YEAR_RE = re.compile(r"(20[2-9]\d)")

def month_range(start_year=2022, start_month=1) -> List[Tuple[int,int]]:
    out = []
    today = datetime.today()
    y, m = start_year, start_month
    while (y < today.year) or (y == today.year and m <= today.month):
        out.append((y, m))
        m += 1
        if m == 13:
            y += 1
            m = 1
    return out

# ── HELPERS ───────────────────────────────────────────────────────────────────
def to_int_or_none(x) -> Optional[int]:
    if x is None:
        return None
    s = str(x).strip().replace(" ", "").replace("\u00A0", "")
    # kill any footnote markers like "123*"
    s = re.sub(r"[^\d\-,\.]", "", s)
    try:
        # integers expected; accept floats that look like "123.0"
        return int(round(float(s)))
    except Exception:
        return None

def is_good_title(title: Optional[str]) -> bool:
    if not title or not title.strip():
        return False
    low = title.lower()
    return any(h in low for h in EXPECTED_TITLE_HINTS)

def read_sheet7_bundle(path: Path) -> Dict[str, Optional[str]]:
    if path.suffix.lower() in [".xlsx", ".xlsm"]:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        ws = wb.worksheets[SHEET_INDEX]
        def cell(addr: str):
            v = ws[addr].value
            return None if v is None else str(v).strip()
        a1 = cell("A1")
        v407 = cell(f"{COL_I}{ROW_407}")
        v408 = cell(f"{COL_I}{ROW_408}")
        label115 = cell(f"{LABEL_COL}{ROW_407}")
        label116 = cell(f"{LABEL_COL}{ROW_408}")
        return dict(title=a1, val407=v407, val408=v408, label115=label115, label116=label116)
    else:
        # .xls via pandas/xlrd
        df = pd.read_excel(path, sheet_name=SHEET_INDEX, header=None)
        def iget(r, c):
            try:
                v = df.iat[r, c]
            except Exception:
                return None
            if pd.isna(v):
                return None
            return str(v).strip()
        a1 = iget(0,0)
        # 0-based indices: I=8, B=1; rows r-1
        v407 = iget(ROW_407-1, 8)
        v408 = iget(ROW_408-1, 8)
        label115 = iget(ROW_407-1, 1)
        label116 = iget(ROW_408-1, 1)
        return dict(title=a1, val407=v407, val408=v408, label115=label115, label116=label116)


def infer_period_from_title(title: str) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    if not title:
        return None, None, None
    t = title.lower()
    year = None
    m = YEAR_RE.search(t)
    if m:
        year = int(m.group(1))

    # normalize dashes
    t = t.replace("—", "-").replace("–", "-")
    # find all possible month tokens (UA + translit)
    def detect_month(word: str) -> Optional[int]:
        for k, v in MONTHS_UA.items():
            if k in word:
                return v
        for k, v in MONTHS_LAT.items():
            if k in word:
                return v
        return None

    # collect tokens that look like months
    tokens = re.split(r"[^a-zA-Zа-яА-ЯїЇіІєЄґҐ]+", t)
    months = [detect_month(w) for w in tokens]
    months = [m for m in months if m]

    if "-" in t or "по" in t or "–" in t:
        # likely a range "січень-березень"
        if len(months) >= 2:
            return year, months[0], months[1]
    # single month like "за січень 2024 року"
    if len(months) >= 1:
        return year, months[0], months[0]
    return year, None, None

def parse_year_month_from_filename(p: Path) -> Tuple[Optional[int], Optional[int]]:
    name = p.stem.lower()
    y = None
    m = YEAR_RE.search(name)
    if m: y = int(m.group(1))
    # try to pick the **last** month token (end-month), since files are Jan–X
    end_month = None
    for token, num in {**MONTHS_LAT, **MONTHS_UA}.items():
        if token in name:
            if end_month is None or name.rfind(token) > name.rfind(list({**MONTHS_LAT, **MONTHS_UA}.keys())[0]):
                end_month = num
    return y, end_month

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    folder = Path(INPUT_DIR)
    files = sorted([p for p in folder.iterdir() if p.suffix.lower() in (".xlsx", ".xls", ".xlsm")])

    cumulative_rows = []
    for p in files:
        try:
            d = read_sheet7_bundle(p)
            title_ok = is_good_title(d["title"])
            y_t, m_start_t, m_end_t = infer_period_from_title(d["title"] or "")
            y_f, m_end_f = parse_year_month_from_filename(p)

            # choose year/end-month preferring title, fallback to filename
            year = y_t or y_f
            end_month = m_end_t or m_end_f
            start_month = m_start_t or 1 if end_month else None  # assume Jan→end_month

            v407 = to_int_or_none(d["val407"])
            v408 = to_int_or_none(d["val408"])

            cumulative_rows.append({
                "file": p.name,
                "year": year,
                "start_month_assumed": start_month if end_month else None,
                "end_month": end_month,
                "sheet7_title": d["title"],
                "title_sanity_ok": title_ok,
                "label_row115_BtoG": d["label115"],
                "label_row116_BtoG": d["label116"],
                "label115_mentions_407": "407" in (d["label115"] or ""),
                "label116_mentions_408": "408" in (d["label116"] or ""),
                "art407_cum": v407,
                "art408_cum": v408,
            })
        except Exception as e:
            cumulative_rows.append({
                "file": p.name, "year": None, "start_month_assumed": None, "end_month": None,
                "sheet7_title": None, "title_sanity_ok": False,
                "label_row115_BtoG": None, "label_row116_BtoG": None,
                "label115_mentions_407": False, "label116_mentions_408": False,
                "art407_cum": None, "art408_cum": None, "error": f"{type(e).__name__}: {e}",
            })

    # write per-file cumulative snapshot
    cum_cols = ["file","year","end_month","start_month_assumed",
                "art407_cum","art408_cum",
                "label_row115_BtoG","label115_mentions_407",
                "label_row116_BtoG","label116_mentions_408",
                "sheet7_title","title_sanity_ok","error"]
    with open(OUTPUT_CUMULATIVE, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cum_cols)
        w.writeheader()
        for r in cumulative_rows:
            if "error" not in r: r["error"] = ""
            w.writerow(r)

    # build per-year cumulative series (choose latest file for each (year,end_month))
    best: Dict[Tuple[int,int], Dict] = {}
    for r in cumulative_rows:
        y, m = r.get("year"), r.get("end_month")
        if y and m:
            key = (y, m)
            # take the row with non-null values; if multiple, prefer the later filename lexicographically
            if key not in best or (r.get("art407_cum") is not None and r.get("art408_cum") is not None and r["file"] > best[key]["file"]):
                best[key] = r

    # create continuous cumulative per month, then difference within each year
    needed = month_range(2022, 1)
    monthly_rows = []
    prev_cum_407: Dict[int, int] = {}
    prev_cum_408: Dict[int, int] = {}

    for (y, m) in needed:
        row = {
            "year": y, "month": m,
            "art407_cum": None, "art408_cum": None,
            "art407_month": None, "art408_month": None,
            "source_file": ""
        }
        if (y, m) in best:
            r = best[(y, m)]
            row["art407_cum"] = r.get("art407_cum")
            row["art408_cum"] = r.get("art408_cum")
            row["source_file"] = r.get("file","")

        # monthly = current cumulative - previous cumulative within the **same year**
        if row["art407_cum"] is not None:
            pm = prev_cum_407.get(y)
            row["art407_month"] = row["art407_cum"] - (pm if pm is not None else 0)
            prev_cum_407[y] = row["art407_cum"]
        if row["art408_cum"] is not None:
            pm = prev_cum_408.get(y)
            row["art408_month"] = row["art408_cum"] - (pm if pm is not None else 0)
            prev_cum_408[y] = row["art408_cum"]

        monthly_rows.append(row)

    # write monthly series
    with open(OUTPUT_MONTHLY, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["year","month","art407_cum","art408_cum","art407_month","art408_month","source_file"])
        w.writeheader()
        for r in monthly_rows:
            w.writerow(r)

    # print coverage based on detected (year,end_month)
    present = {(r["year"], r["end_month"]) for r in cumulative_rows if r.get("year") and r.get("end_month")}
    missing = [f"{y}-{m:02d}" for (y, m) in needed if (y, m) not in present]
    print(f"Wrote: {OUTPUT_CUMULATIVE}  ({len(cumulative_rows)} files)")
    print(f"Wrote: {OUTPUT_MONTHLY}    ({len(monthly_rows)} rows)")
    if missing:
        print("\nMissing months (no cumulative snapshot found for these YYYY-MM):")
        print(", ".join(missing))
    else:
        print("\nCoverage OK: snapshot found for every month from 2022-01 to current month.")

if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        sys.exit(1)
